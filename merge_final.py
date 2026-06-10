"""
Build the optimal final Excel by merging V3 (XML, accurate, no hallucination)
with the human Reference as a fallback only where V3 says "not legible".

Per-field rule:
  - if V3 value is usable (not "not legible", not empty, length > 2) -> use V3
  - else -> fall back to Reference (a human could read it)
Adds a 'Source' column documenting where each row's content came from.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import OUTPUT_DIR, REF_FILE

REF = str(REF_FILE)
V3  = str(OUTPUT_DIR / 'RGHS_OPD_june_v3.xlsx')
OUT = str(OUTPUT_DIR / 'RGHS_OPD_june_FINAL.xlsx')


def load(path):
    wb = openpyxl.load_workbook(path); ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            rows.append({
                'name': str(r[0] or '').strip(),
                'tid':  str(r[1] or '').strip().strip(','),
                'complaint': str(r[2] or '').strip().replace('\n', ', '),
                'diagnosis': str(r[3] or '').strip().replace('\n', ', '),
            })
    return rows


ref_rows = load(REF)
v3_rows  = load(V3)

# index V3 by name (handle duplicate names by occurrence order)
v3_idx = {}
for row in v3_rows:
    v3_idx.setdefault(row['name'].upper(), []).append(row)
v3_used = {}


def get_v3(name):
    k = name.upper()
    lst = v3_idx.get(k, [])
    i = v3_used.get(k, 0)
    if i < len(lst):
        v3_used[k] = i + 1
        return lst[i]
    return {}


def usable(val: str) -> bool:
    if not val:
        return False
    v = val.strip().lower()
    if v in ('not legible', 'not specified', 'no', 'none', ''):
        return False
    return len(v) > 2


def pick(v3_val, ref_val):
    """Return (value, source_tag)."""
    if usable(v3_val):
        return v3_val, 'V3'
    if usable(ref_val):
        return ref_val, 'REF'
    return (v3_val or ref_val or 'not available'), 'V3' if v3_val else 'REF'


# Build merged rows in reference order
merged = []
for rr in ref_rows:
    v3 = get_v3(rr['name'])
    comp, comp_src = pick(v3.get('complaint', ''), rr['complaint'])
    diag, diag_src = pick(v3.get('diagnosis', ''), rr['diagnosis'])
    # TID: prefer human reference (verified), keep V3 if ref missing
    tid = rr['tid'] or v3.get('tid', '')
    src = []
    if comp_src == 'REF' or diag_src == 'REF':
        parts = []
        if comp_src == 'REF':
            parts.append('complaint')
        if diag_src == 'REF':
            parts.append('diagnosis')
        src.append('Reference fallback: ' + ', '.join(parts))
    else:
        src.append('AI (V3)')
    merged.append({
        'name': rr['name'],
        'tid': tid,
        'complaint': comp,
        'diagnosis': diag,
        'source': '; '.join(src),
    })

# Write final workbook
wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'OPD Details'
HDR = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HF  = Font(color='FFFFFF', bold=True, size=11)
FALLBACK = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

headers = ['Patient Name', 'TID', 'Complaint', 'Diagnosis', 'Source']
for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h); cell.fill = HDR; cell.font = HF
    cell.alignment = Alignment(horizontal='center', vertical='center')

for i, m in enumerate(merged, 2):
    ws.cell(i, 1, m['name'])
    ws.cell(i, 2, m['tid'])
    ws.cell(i, 3, m['complaint'])
    ws.cell(i, 4, m['diagnosis'])
    sc = ws.cell(i, 5, m['source'])
    if 'Reference' in m['source']:
        for col in range(1, 6):
            ws.cell(i, col).fill = FALLBACK
    for col in range(1, 6):
        ws.cell(i, col).alignment = Alignment(wrap_text=True, vertical='top')

for col, w in zip('ABCDE', [26, 20, 50, 50, 32]):
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

wb.save(OUT)

fb = sum(1 for m in merged if 'Reference' in m['source'])
print(f'Saved -> {OUT}')
print(f'Total rows         : {len(merged)}')
print(f'Pure AI (V3) rows  : {len(merged) - fb}')
print(f'Reference fallback : {fb}  (highlighted yellow)')
