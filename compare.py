"""
Compare extracted v2 Excel vs reference Excel shared by doctor.
Match on TID, compare Patient Name / Complaint / Diagnosis.
Output a detailed comparison Excel with colour-coded differences.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from difflib import SequenceMatcher
from config import OUTPUT_DIR, REF_FILE as _REF_FILE

REF_FILE  = str(_REF_FILE)
EXT_FILE  = str(OUTPUT_DIR / 'RGHS_OPD_june_v2.xlsx')
OUT_FILE  = str(OUTPUT_DIR / 'comparison_report.xlsx')

# ── Colours ───────────────────────────────────────────────────────────────────
GREEN  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # exact match
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # partial match
RED    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # mismatch / missing
BLUE   = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # header
GREY   = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # not in ref


def normalise(text: str) -> str:
    if not text:
        return ''
    text = str(text).lower().strip()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'[,;]+', ',', text)
    return text.strip(' ,')


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalise(a), normalise(b)).ratio()


def classify(a: str, b: str):
    """Return (fill, label) based on similarity between two strings."""
    if not a and not b:
        return GREEN, 'BOTH EMPTY'
    if not a:
        return RED, 'MISSING IN EXTRACTED'
    if not b:
        return RED, 'MISSING IN REFERENCE'
    score = similarity(a, b)
    if score >= 0.85:
        return GREEN, f'MATCH ({score:.0%})'
    if score >= 0.45:
        return YELLOW, f'PARTIAL ({score:.0%})'
    return RED, f'MISMATCH ({score:.0%})'


# ── Load files ────────────────────────────────────────────────────────────────
def load_by_tid(path, name_col=0, tid_col=1, complaint_col=2, diag_col=3):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = str(row[tid_col] or '').strip().strip(',').strip()
        if not tid or tid == 'None':
            continue
        rows[tid] = {
            'name':      str(row[name_col] or '').strip(),
            'complaint': str(row[complaint_col] or '').strip(),
            'diagnosis': str(row[diag_col] or '').strip(),
        }
    return rows


ref = load_by_tid(REF_FILE)
ext = load_by_tid(EXT_FILE)

print(f'Reference records : {len(ref)}')
print(f'Extracted records : {len(ext)}')

all_tids = sorted(set(ref) | set(ext))

# ── Build comparison Excel ────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Comparison'

thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)

headers = [
    'TID',
    'REF — Patient Name', 'EXT — Patient Name', 'Name Match',
    'REF — Complaint',    'EXT — Complaint',    'Complaint Match',
    'REF — Diagnosis',    'EXT — Diagnosis',    'Diagnosis Match',
    'Status',
]
for col, h in enumerate(headers, 1):
    cell = ws_out.cell(row=1, column=col, value=h)
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.font = Font(color='FFFFFF', bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin

# ── Per-field score tracking ──────────────────────────────────────────────────
scores = {'name': [], 'complaint': [], 'diagnosis': []}
statuses = {'EXACT': 0, 'PARTIAL': 0, 'MISMATCH': 0, 'NOT IN REF': 0, 'NOT IN EXT': 0}

for row_idx, tid in enumerate(all_tids, 2):
    r = ref.get(tid, {})
    e = ext.get(tid, {})

    if not r:
        status = 'NOT IN REF'
        status_fill = GREY
    elif not e:
        status = 'NOT IN EXT'
        status_fill = RED
    else:
        name_score      = similarity(r['name'],      e['name'])
        complaint_score = similarity(r['complaint'],  e['complaint'])
        diag_score      = similarity(r['diagnosis'],  e['diagnosis'])
        avg = (name_score + complaint_score + diag_score) / 3
        scores['name'].append(name_score)
        scores['complaint'].append(complaint_score)
        scores['diagnosis'].append(diag_score)

        if avg >= 0.85:
            status, status_fill = 'EXACT', GREEN
            statuses['EXACT'] += 1
        elif avg >= 0.45:
            status, status_fill = 'PARTIAL', YELLOW
            statuses['PARTIAL'] += 1
        else:
            status, status_fill = 'MISMATCH', RED
            statuses['MISMATCH'] += 1

    if not r:
        statuses['NOT IN REF'] += 1
    elif not e:
        statuses['NOT IN EXT'] += 1

    name_fill,  name_label  = classify(r.get('name',''),      e.get('name',''))
    comp_fill,  comp_label  = classify(r.get('complaint',''), e.get('complaint',''))
    diag_fill,  diag_label  = classify(r.get('diagnosis',''), e.get('diagnosis',''))

    values = [
        (tid,                   None),
        (r.get('name',''),      name_fill),
        (e.get('name',''),      name_fill),
        (name_label,            name_fill),
        (r.get('complaint',''), comp_fill),
        (e.get('complaint',''), comp_fill),
        (comp_label,            comp_fill),
        (r.get('diagnosis',''), diag_fill),
        (e.get('diagnosis',''), diag_fill),
        (diag_label,            diag_fill),
        (status,                status_fill),
    ]
    for col, (val, fill) in enumerate(values, 1):
        cell = ws_out.cell(row=row_idx, column=col, value=val)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = Font(size=9)
        cell.border = thin

# ── Column widths ─────────────────────────────────────────────────────────────
widths = [20, 22, 22, 18, 40, 40, 18, 40, 40, 18, 14]
for col, w in enumerate(widths, 1):
    ws_out.column_dimensions[ws_out.cell(1, col).column_letter].width = w
ws_out.row_dimensions[1].height = 30
ws_out.freeze_panes = 'B2'

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws_s = wb_out.create_sheet('Summary')

def avg_pct(lst):
    return f'{sum(lst)/len(lst)*100:.1f}%' if lst else 'N/A'

summary_rows = [
    ('OVERALL', ''),
    ('Reference records',           len(ref)),
    ('Extracted records',           len(ext)),
    ('Matched by TID',              len([t for t in all_tids if t in ref and t in ext])),
    ('In extracted but not in ref', statuses['NOT IN REF']),
    ('In ref but not extracted',    statuses['NOT IN EXT']),
    ('', ''),
    ('MATCH QUALITY (matched records only)', ''),
    ('Exact matches (avg ≥85%)',    statuses['EXACT']),
    ('Partial matches (45–85%)',    statuses['PARTIAL']),
    ('Mismatches (<45%)',           statuses['MISMATCH']),
    ('', ''),
    ('FIELD ACCURACY', ''),
    ('Patient Name avg similarity',  avg_pct(scores['name'])),
    ('Complaint avg similarity',     avg_pct(scores['complaint'])),
    ('Diagnosis avg similarity',     avg_pct(scores['diagnosis'])),
]

for i, (label, value) in enumerate(summary_rows, 1):
    ws_s.cell(i, 1, label).font = Font(bold=bool(value == '' or label.isupper() or label == label.upper()), size=11)
    ws_s.cell(i, 2, value).font = Font(size=11)

ws_s.column_dimensions['A'].width = 42
ws_s.column_dimensions['B'].width = 20

wb_out.save(OUT_FILE)
print(f'\nSaved → {OUT_FILE}')
print(f'\n=== SUMMARY ===')
for label, value in summary_rows:
    if label:
        print(f'  {label:<42} {value}')
