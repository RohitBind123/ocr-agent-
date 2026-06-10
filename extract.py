"""
RGHS OPD Prescription Extractor
- Async parallel Gemini calls (asyncio.gather + semaphore)
- Chain-of-thought system prompt for clean clinical output
- Concurrency capped at 5 to stay within Gemini rate limits
"""

import asyncio
import base64
import io
import json
import os
import re
import time
from pathlib import Path

import google.genai as genai
from google.genai import types
from pdf2image import convert_from_path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from prompt import SYSTEM_PROMPT, EXTRACTION_PROMPT
from config import GEMINI_API_KEY, PDF_DIR, OUTPUT_DIR

# ── Config ────────────────────────────────────────────────────────────────────
OUTPUT_FILE = OUTPUT_DIR / 'RGHS_OPD_june_v3.xlsx'

CONCURRENCY    = 5      # max parallel Gemini calls at once
THINKING_BUDGET = 3000  # LOW budget — research shows high budget increases hallucination on transcription
DPI            = 200    # higher render quality for better handwriting legibility


# ── Helpers ───────────────────────────────────────────────────────────────────
def pdf_to_jpeg_bytes(path: Path) -> bytes:
    pages = convert_from_path(str(path), dpi=DPI)
    buf = io.BytesIO()
    pages[0].save(buf, format='JPEG', quality=92)
    return buf.getvalue()


def parse_json(raw: str) -> dict:
    raw = raw.strip()
    raw = re.sub(r'^```[a-z]*\n?', '', raw)
    raw = re.sub(r'\n?```$', '', raw)
    match = re.search(r'\{.*\}', raw, re.DOTALL)
    if match:
        raw = match.group(0)
    return json.loads(raw)


# ── Async extraction ──────────────────────────────────────────────────────────
async def extract_one(
    client: genai.Client,
    sem: asyncio.Semaphore,
    idx: int,
    total: int,
    fname: str,
) -> dict:
    path = PDF_DIR / fname
    async with sem:
        print(f"[{idx:02d}/{total}] → {fname}")
        try:
            # PDF → JPEG in thread pool (blocking I/O)
            loop = asyncio.get_event_loop()
            img_bytes = await loop.run_in_executor(None, pdf_to_jpeg_bytes, path)
            b64 = base64.b64encode(img_bytes).decode()

            response = await client.aio.models.generate_content(
                model='gemini-2.5-flash',
                contents=[
                    types.Content(
                        role='user',
                        parts=[
                            types.Part(
                                inline_data=types.Blob(
                                    mime_type='image/jpeg',
                                    data=b64,
                                )
                            ),
                            types.Part(text=EXTRACTION_PROMPT),
                        ],
                    )
                ],
                config=types.GenerateContentConfig(
                    system_instruction=SYSTEM_PROMPT,
                    thinking_config=types.ThinkingConfig(
                        thinking_budget=THINKING_BUDGET
                    ),
                    temperature=0.1,
                ),
            )

            data = parse_json(response.text)
            data['_source'] = fname
            print(f"    ✓  {data.get('patient_name')} | {str(data.get('tid',''))[:16]}")
            print(f"       Complaint : {data.get('complaint','')[:90]}")
            print(f"       Diagnosis : {data.get('diagnosis','')[:90]}")
            return data

        except Exception as exc:
            print(f"    ✗  ERROR: {exc}")
            return {
                'patient_name': fname.replace(' - P.pdf', ''),
                'tid': 'ERROR',
                'complaint': str(exc),
                'diagnosis': '',
                '_source': fname,
            }


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(rows: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OPD Details'

    headers = ['Patient Name', 'TID', 'Complaint', 'Diagnosis']
    h_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    h_font = Font(color='FFFFFF', bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row.get('patient_name', ''))
        ws.cell(row=i, column=2, value=str(row.get('tid', '')).strip(',').strip())
        ws.cell(row=i, column=3, value=row.get('complaint', ''))
        ws.cell(row=i, column=4, value=row.get('diagnosis', ''))
        for col in range(1, 5):
            ws.cell(row=i, column=col).alignment = Alignment(
                wrap_text=True, vertical='top'
            )

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 55
    ws.row_dimensions[1].height = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ── Main ──────────────────────────────────────────────────────────────────────
async def main() -> None:
    client = genai.Client(api_key=GEMINI_API_KEY)
    sem    = asyncio.Semaphore(CONCURRENCY)

    p_files = sorted(f for f in os.listdir(PDF_DIR) if f.endswith('- P.pdf'))
    total   = len(p_files)
    print(f"Found {total} prescription PDFs  |  concurrency={CONCURRENCY}\n")

    t0 = time.time()

    # Fan out all extractions in parallel, capped by semaphore
    tasks = [
        extract_one(client, sem, idx, total, fname)
        for idx, fname in enumerate(p_files, 1)
    ]
    results = await asyncio.gather(*tasks)

    # Preserve alphabetical order (gather returns in task-creation order)
    rows = [r for r in results if r]

    build_excel(rows, OUTPUT_FILE)

    elapsed = time.time() - t0
    errors  = [r for r in rows if r.get('tid') == 'ERROR']
    print(f"\n{'='*65}")
    print(f"Done in {elapsed:.1f}s  |  {len(rows)} rows  |  {len(errors)} errors")
    print(f"Output → {OUTPUT_FILE}")
    if errors:
        print(f"Failed files: {[e['_source'] for e in errors]}")


if __name__ == '__main__':
    asyncio.run(main())
