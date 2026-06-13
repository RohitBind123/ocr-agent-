"""
Excel writers for the case pipeline:
  - write_extraction_excel : one model's output in the ground-truth 5-column format
  - write_bakeoff_report   : multi-sheet model comparison (summary + per-model detail)
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.analysis.accuracy_judge import SCORED_FIELDS

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
GOOD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BAD = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

GT_HEADERS = ["Patient name", "Consultant Name", "Complain", "Diagnosis", "Duration"]


def _style_header(ws: Worksheet, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 22


def _display_name(rec: dict) -> str:
    tid = (rec.get("tid") or rec.get("folder_tid") or "").strip()
    name = (rec.get("patient_name") or rec.get("folder_name") or "").strip()
    return f"{tid} {name}".strip()


def write_extraction_excel(rows: list[dict], output_path: Path) -> None:
    """Write one model's extractions in the ground-truth 5-column layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted"
    _style_header(ws, GT_HEADERS)
    for i, rec in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=_display_name(rec))
        ws.cell(row=i, column=2, value=rec.get("consultant_name", ""))
        ws.cell(row=i, column=3, value=rec.get("complaint", ""))
        ws.cell(row=i, column=4, value=rec.get("diagnosis", ""))
        ws.cell(row=i, column=5, value=rec.get("duration", ""))
        for col in range(1, 6):
            ws.cell(row=i, column=col).alignment = WRAP_TOP
    for col, width in zip("ABCDE", (30, 22, 50, 50, 14)):
        ws.column_dimensions[col].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def _score_fill(score: float) -> PatternFill | None:
    if score >= 0.99:
        return GOOD
    if score >= 0.5:
        return WARN
    return BAD


def _write_summary(ws: Worksheet, summary_rows: list[dict]) -> None:
    headers = ["Rank", "Model", "Overall %"] + [f"{f} %" for f in SCORED_FIELDS] + [
        "Avg latency (s)", "Errors", "Cases",
    ]
    _style_header(ws, headers)
    for i, row in enumerate(summary_rows, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=row["model"])
        c = ws.cell(row=i, column=3, value=round(row["overall"] * 100, 1))
        c.fill = _score_fill(row["overall"])
        for j, f in enumerate(SCORED_FIELDS, 4):
            fc = ws.cell(row=i, column=j, value=round(row["fields"][f] * 100, 1))
            fc.fill = _score_fill(row["fields"][f])
        ws.cell(row=i, column=4 + len(SCORED_FIELDS), value=row["avg_latency"])
        ws.cell(row=i, column=5 + len(SCORED_FIELDS), value=row["errors"])
        ws.cell(row=i, column=6 + len(SCORED_FIELDS), value=row["n"])
    for col, width in zip("ABCDEFGHIJK", (5, 26, 10, 14, 16, 12, 12, 12, 14, 8, 7)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def _write_model_detail(
    ws: Worksheet, extractions: list[dict], judged: dict[str, dict], gt_by_case: dict
) -> None:
    headers = [
        "Case", "Field", "Ground Truth", "Extracted", "Verdict", "Score", "Sim", "Reason",
    ]
    _style_header(ws, headers)
    r = 2
    for rec in extractions:
        cid = rec["case_id"]
        gt = gt_by_case.get(cid)
        jcase = judged.get(cid, {})
        jfields = jcase.get("fields", {})
        for f in SCORED_FIELDS:
            jf = jfields.get(f, {})
            ws.cell(row=r, column=1, value=cid if f == SCORED_FIELDS[0] else "")
            ws.cell(row=r, column=2, value=f)
            ws.cell(row=r, column=3, value=getattr(gt, f, "") if gt else "")
            ws.cell(row=r, column=4, value=rec.get(f, ""))
            vc = ws.cell(row=r, column=5, value=jf.get("verdict", ""))
            sc = ws.cell(row=r, column=6, value=jf.get("score", ""))
            fill = _score_fill(float(jf.get("score", 0) or 0))
            if fill:
                vc.fill = fill
                sc.fill = fill
            ws.cell(row=r, column=7, value=jf.get("similarity", ""))
            ws.cell(row=r, column=8, value=jf.get("reason", ""))
            for col in (3, 4, 8):
                ws.cell(row=r, column=col).alignment = WRAP_TOP
            r += 1
    for col, width in zip("ABCDEFGH", (30, 16, 42, 42, 9, 7, 7, 44)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def write_bakeoff_report(
    bundles: list[dict],
    gt_by_case: dict,
    output_path: Path,
) -> None:
    """
    bundles: list of per-model dicts, each:
      {model, extractions:[...], judged:{case_id:judgement}, agg:{overall,fields,n},
       avg_latency, errors}
    Ranked by overall accuracy descending.
    """
    bundles = sorted(bundles, key=lambda b: b["agg"]["overall"], reverse=True)
    wb = openpyxl.Workbook()

    summary_rows = [
        {
            "model": b["model"],
            "overall": b["agg"]["overall"],
            "fields": b["agg"]["fields"],
            "avg_latency": b["avg_latency"],
            "errors": b["errors"],
            "n": b["agg"]["n"],
        }
        for b in bundles
    ]
    _write_summary(wb.active, summary_rows)
    wb.active.title = "Summary"

    for b in bundles:
        safe = b["model"].replace("gemini-", "").replace("-preview", "")[:28]
        ws = wb.create_sheet(title=safe)
        _write_model_detail(ws, b["extractions"], b["judged"], gt_by_case)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
