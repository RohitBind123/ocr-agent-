"""Shared Excel style constants and the common OPD workbook writer."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP_TOP    = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")


def write_opd_excel(rows: list[dict], output_path: Path) -> None:
    """Write a 4-column OPD Excel (Name / TID / Complaint / Diagnosis)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OPD Details"

    headers = ["Patient Name", "TID", "Complaint", "Diagnosis"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for i, row in enumerate(rows, 2):
        # support both key names used across pipeline stages
        name = row.get("patient_name") or row.get("name", "")
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=str(row.get("tid", "")).strip(",").strip())
        ws.cell(row=i, column=3, value=row.get("complaint", ""))
        ws.cell(row=i, column=4, value=row.get("diagnosis", ""))
        for col in range(1, 5):
            ws.cell(row=i, column=col).alignment = WRAP_TOP

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 55
    ws.row_dimensions[1].height = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
