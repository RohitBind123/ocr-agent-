"""
ReAct retry agent for illegible RGHS prescription fields.

Loop: the model sees the current image, REASONS (Thought), and either requests an
ACTION (zoom / enhance / isolate_ink on a named region) or returns a FINAL answer.
Each action produces a new image fed back to the model. Max N iterations.

This lets the model actively investigate hard handwriting instead of giving up.
"""
import asyncio
import json
import os

import google.genai as genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.utils import image_tools as IT
from src.utils.parsing import parse_json
from src.config import GEMINI_API_KEY, PDF_DIR, OUTPUT_DIR
from src.prompts.react import REACT_SYSTEM

PDF_DIR_STR = str(PDF_DIR)
V3  = str(OUTPUT_DIR / "RGHS_OPD_june_v3.xlsx")
OUT = str(OUTPUT_DIR / "RGHS_OPD_june_v4_react.xlsx")
LOG = str(OUTPUT_DIR / "react_trace.txt")

MAX_ITERS   = 4
CONCURRENCY = 3


async def react_one(client, sem, name, fields, base_img, trace):
    """Run the ReAct loop for one record. Returns dict with possibly-updated fields."""
    async with sem:
        cur_img = base_img
        ask = (
            f"Fields marked illegible to re-read: {', '.join(fields)}. "
            f"Investigate the relevant region(s) and decipher them. Start now."
        )

        lines = [f"\n{'='*60}\n{name}  | re-read: {fields}"]
        contents = [types.Content(role="user", parts=[
            types.Part(inline_data=types.Blob(mime_type="image/jpeg", data=IT.to_jpeg_b64(cur_img))),
            types.Part(text=ask),
        ])]

        result = {"complaint": None, "diagnosis": None, "confidence": "low"}

        for it in range(1, MAX_ITERS + 1):
            try:
                resp = await client.aio.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=contents,
                    config=types.GenerateContentConfig(
                        system_instruction=REACT_SYSTEM,
                        thinking_config=types.ThinkingConfig(thinking_budget=2500),
                        temperature=0.1),
                )
                step = parse_json(resp.text)
            except Exception as e:
                lines.append(f"  iter {it}: ERROR {e}")
                break

            thought = step.get("thought", "")[:100]
            action = step.get("action")
            lines.append(f"  iter {it}: [{action}] {thought}")

            if action == "final":
                result["complaint"] = step.get("complaint")
                result["diagnosis"] = step.get("diagnosis")
                result["confidence"] = step.get("confidence", "low")
                lines.append(f"    FINAL ({result['confidence']}): "
                             f"comp={result['complaint']!r} diag={result['diagnosis']!r}")
                break

            # execute image action
            new_img, obs = IT.apply_action(base_img, step)
            cur_img = new_img
            # append assistant turn + new observation image
            contents.append(types.Content(role="model", parts=[types.Part(text=json.dumps(step))]))
            contents.append(types.Content(role="user", parts=[
                types.Part(inline_data=types.Blob(mime_type="image/jpeg", data=IT.to_jpeg_b64(cur_img))),
                types.Part(text=f"Observation: {obs} Continue, or give final answer."),
            ]))
        else:
            lines.append("    (max iters reached, no final)")

        trace.append("\n".join(lines))
        print(f"  ✓ {name:<22} conf={result['confidence']:<7} "
              f"diag={str(result['diagnosis'])[:50]}")
        return name, fields, result


async def main():
    client = genai.Client(api_key=GEMINI_API_KEY)
    sem = asyncio.Semaphore(CONCURRENCY)

    # load v3
    wb = openpyxl.load_workbook(V3); ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            rows.append({"name": str(r[0]).strip(), "tid": str(r[1] or "").strip(),
                         "complaint": str(r[2] or "").strip(), "diagnosis": str(r[3] or "").strip()})

    # find not-legible (dedupe by name+order to handle dup names)
    targets = []
    for i, row in enumerate(rows):
        comp, diag = row["complaint"].lower(), row["diagnosis"].lower()
        f = []
        if "not legible" in comp or comp in ("no", "none", ""): f.append("complaint")
        if "not legible" in diag or diag in ("not specified", "none", ""): f.append("diagnosis")
        if f:
            targets.append((i, row["name"], f))

    print(f"ReAct retry on {len(targets)} records (max {MAX_ITERS} iters each)\n")

    # render PDFs (blocking) up front
    loop = asyncio.get_event_loop()
    trace = []
    tasks = []
    seen = {}
    for i, name, fields in targets:
        # pick correct PDF for duplicate names by occurrence
        key = name.upper()
        pdfs = sorted(f for f in os.listdir(PDF_DIR_STR) if f.endswith("- P.pdf")
                      and f.split(" - ")[0].upper().replace(" ", "").startswith(key.replace(" ", "")[:8]))
        occ = seen.get(key, 0); seen[key] = occ + 1
        pdf = pdfs[occ] if occ < len(pdfs) else (pdfs[0] if pdfs else None)
        if not pdf:
            print(f"  ! no PDF for {name}")
            continue
        img = await loop.run_in_executor(None, IT.render_page, os.path.join(PDF_DIR_STR, pdf), 300)
        tasks.append((i, react_one(client, sem, name, fields, img, trace)))

    results = await asyncio.gather(*[t for _, t in tasks])
    idx_map = [i for i, _ in tasks]

    # apply improvements
    improved = 0
    for (i), (name, fields, res) in zip(idx_map, results):
        for fld in fields:
            val = res.get(fld)
            if val and val != "keep" and "not legible" not in str(val).lower():
                rows[i][fld] = val
                improved += 1

    # write v4
    wbo = openpyxl.Workbook(); wso = wbo.active; wso.title = "OPD Details"
    HDR = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for c, h in enumerate(["Patient Name", "TID", "Complaint", "Diagnosis"], 1):
        cell = wso.cell(1, c, h); cell.fill = HDR
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(rows, 2):
        wso.cell(r, 1, row["name"]); wso.cell(r, 2, row["tid"])
        wso.cell(r, 3, row["complaint"]); wso.cell(r, 4, row["diagnosis"])
        for c in range(1, 5):
            wso.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCD", [26, 20, 50, 50]):
        wso.column_dimensions[col].width = w
    wbo.save(OUT)

    with open(LOG, "w") as f:
        f.write("\n".join(trace))

    print(f"\n{'='*60}")
    print(f"ReAct done. Improved {improved} fields across {len(targets)} records.")
    print(f"Output -> {OUT}")
    print(f"Trace  -> {LOG}")


if __name__ == "__main__":
    asyncio.run(main())
