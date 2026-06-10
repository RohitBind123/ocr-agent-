"""3-way comparison: Reference vs v2 (CoT) vs v3 (XML). Side-by-side per patient."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import OUTPUT_DIR, REF_FILE

REF = str(REF_FILE)
V2  = str(OUTPUT_DIR / 'RGHS_OPD_june_v2.xlsx')
V3  = str(OUTPUT_DIR / 'RGHS_OPD_june_v3.xlsx')
OUT = str(OUTPUT_DIR / 'comparison_3way.xlsx')


def load(path):
    wb = openpyxl.load_workbook(path); ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            rows.append({
                'name': str(r[0] or '').strip(),
                'tid':  str(r[1] or '').strip().strip(','),
                'complaint': str(r[2] or '').strip().replace('\n', ', '),
                'diagnosis': str(r[3] or '').strip().replace('\n', ', '),
            })
    return rows


ref, v2, v3 = load(REF), load(V2), load(V3)

# index v2/v3 by normalized name (handle dup names by order)
def name_key(n): return n.strip().upper()

def build_index(rows):
    idx = {}
    for row in rows:
        idx.setdefault(name_key(row['name']), []).append(row)
    return idx

v2i, v3i = build_index(v2), build_index(v3)
v2_used, v3_used = {}, {}

def match(name, idx, used):
    k = name_key(name)
    lst = idx.get(k, [])
    i = used.get(k, 0)
    if i < len(lst):
        used[k] = i + 1
        return lst[i]
    return {}

# count not-legible per version
def count_nl(rows):
    c = sum(1 for r in rows if 'not legible' in r['complaint'].lower())
    d = sum(1 for r in rows if 'not legible' in r['diagnosis'].lower())
    return c, d

v2_nl = count_nl(v2); v3_nl = count_nl(v3)

# Build comparison workbook
wb = openpyxl.Workbook(); ws = wb.active; ws.title = '3-Way'
thin = Border(*[Side(style='thin')]*4)
HDR = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
REFC = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
V2C  = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
V3C  = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

headers = ['Patient', 'Field', 'REFERENCE (human)', 'V2 (CoT, budget 8k)', 'V3 (XML, budget 3k)']
for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h); cell.fill = HDR
    cell.font = Font(color='FFFFFF', bold=True); cell.border = thin
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row = 2
for rrow in ref:
    e2 = match(rrow['name'], v2i, v2_used)
    e3 = match(rrow['name'], v3i, v3_used)
    for field in ('complaint', 'diagnosis'):
        ws.cell(row, 1, rrow['name']).border = thin
        ws.cell(row, 2, field.capitalize()).border = thin
        c3r = ws.cell(row, 3, rrow.get(field, '')); c3r.fill = REFC; c3r.border = thin
        c2  = ws.cell(row, 4, e2.get(field, '')); c2.fill = V2C; c2.border = thin
        c3  = ws.cell(row, 5, e3.get(field, '')); c3.fill = V3C; c3.border = thin
        for cc in (1, 2, 3, 4, 5):
            ws.cell(row, cc).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row, cc).font = Font(size=9)
        row += 1

for col, w in zip('ABCDE', [20, 12, 45, 45, 45]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'

# summary sheet
s = wb.create_sheet('Summary')
data = [
    ('Total records', len(ref), len(v2), len(v3)),
    ('', '', '', ''),
    ('"not legible" — Complaint', '-', v2_nl[0], v3_nl[0]),
    ('"not legible" — Diagnosis', '-', v2_nl[1], v3_nl[1]),
    ('', '', '', ''),
    ('Source', 'REFERENCE', 'V2', 'V3'),
]
s.cell(1, 1, 'Metric').font = Font(bold=True)
for c, h in enumerate(['Metric', 'REF', 'V2', 'V3'], 1):
    s.cell(1, c, h).font = Font(bold=True)
for i, rowdata in enumerate(data, 2):
    for c, val in enumerate(rowdata, 1):
        s.cell(i, c, val)
for col, w in zip('ABCD', [32, 14, 14, 14]):
    s.column_dimensions[col].width = w

wb.save(OUT)
print(f'Saved -> {OUT}')
print(f'\nNot legible counts:')
print(f'  V2: complaint={v2_nl[0]}, diagnosis={v2_nl[1]}')
print(f'  V3: complaint={v3_nl[0]}, diagnosis={v3_nl[1]}')
