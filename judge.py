"""
Clinical judge: compare V3 (AI) vs Reference (human) on CONTEXTUAL medical equivalence,
not word-by-word. Uses Gemini 2.5 Flash as an impartial medical adjudicator.
Parallel async calls.
"""
import asyncio, json, re, os
import google.genai as genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import GEMINI_API_KEY, OUTPUT_DIR, REF_FILE

REF = str(REF_FILE)
V3  = str(OUTPUT_DIR / 'RGHS_OPD_june_v3.xlsx')
OUT = str(OUTPUT_DIR / 'judge_report.xlsx')
CONCURRENCY = 5

JUDGE_SYSTEM = """<role>You are a senior physician adjudicating two transcriptions of the same
handwritten OPD prescription: one by a human (REFERENCE), one by an AI (AI). You judge whether
they convey the same CLINICAL MEANING — not whether the words match.</role>

<verdicts>
- EQUIVALENT: same clinical meaning (e.g. "HTN" vs "Hypertension"; "follow up DM" vs "Type 2 Diabetes Mellitus follow-up")
- AI_RICHER: AI is correct AND adds valid extra detail (duration, laterality, real comorbidities)
- REF_RICHER: Reference is correct AND has detail the AI missed
- AI_WRONG: AI states something clinically different/incorrect vs reference
- REF_WRONG: Reference looks wrong and AI looks more plausible
- AI_GAVEUP: AI returned "not legible" / empty while reference has a value
- BOTH_UNCLEAR: neither is usable
</verdicts>

<output>Return ONLY raw JSON: {"complaint_verdict":"...","diagnosis_verdict":"...","note":"<=12 words"}</output>"""


def load(path):
    wb = openpyxl.load_workbook(path); ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            rows.append({'name': str(r[0]).strip(),
                         'complaint': str(r[2] or '').strip().replace('\n', ', '),
                         'diagnosis': str(r[3] or '').strip().replace('\n', ', ')})
    return rows


ref_rows = load(REF)
v3_rows = load(V3)
v3_idx = {}
for r in v3_rows:
    v3_idx.setdefault(r['name'].upper(), []).append(r)
used = {}
def get_v3(name):
    k = name.upper(); lst = v3_idx.get(k, []); i = used.get(k, 0)
    if i < len(lst): used[k] = i+1; return lst[i]
    return {}


def parse(raw):
    raw = re.sub(r'^```[a-z]*\n?', '', raw.strip()); raw = re.sub(r'\n?```$', '', raw)
    m = re.search(r'\{.*\}', raw, re.DOTALL)
    return json.loads(m.group(0) if m else raw)


async def judge_one(client, sem, idx, total, rr):
    v3 = get_v3(rr['name'])
    async with sem:
        prompt = f"""Patient: {rr['name']}

COMPLAINT:
  REFERENCE: {rr['complaint'] or '(empty)'}
  AI:        {v3.get('complaint','') or '(empty)'}

DIAGNOSIS:
  REFERENCE: {rr['diagnosis'] or '(empty)'}
  AI:        {v3.get('diagnosis','') or '(empty)'}

Adjudicate each field. Return JSON only."""
        try:
            resp = await client.aio.models.generate_content(
                model='gemini-2.5-flash',
                contents=[types.Content(role='user', parts=[types.Part(text=prompt)])],
                config=types.GenerateContentConfig(
                    system_instruction=JUDGE_SYSTEM,
                    thinking_config=types.ThinkingConfig(thinking_budget=1500),
                    temperature=0.0),
            )
            v = parse(resp.text)
            v['name'] = rr['name']
            v['ref_d'] = rr['diagnosis']; v['ai_d'] = v3.get('diagnosis','')
            print(f"[{idx:02d}/{total}] {rr['name'][:20]:<20} C:{v.get('complaint_verdict','?'):<12} D:{v.get('diagnosis_verdict','?')}")
            return v
        except Exception as e:
            print(f"[{idx:02d}/{total}] {rr['name']} ERROR {e}")
            return {'name': rr['name'], 'complaint_verdict': 'ERROR', 'diagnosis_verdict': 'ERROR', 'note': str(e)[:40], 'ref_d':'', 'ai_d':''}


async def main():
    client = genai.Client(api_key=GEMINI_API_KEY)
    sem = asyncio.Semaphore(CONCURRENCY)
    total = len(ref_rows)
    results = await asyncio.gather(*[judge_one(client, sem, i, total, rr) for i, rr in enumerate(ref_rows, 1)])

    # tally
    from collections import Counter
    cv = Counter(r['complaint_verdict'] for r in results)
    dv = Counter(r['diagnosis_verdict'] for r in results)

    # write report
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Judge'
    HDR = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    colors = {
        'EQUIVALENT': 'C6EFCE', 'AI_RICHER': 'BDD7EE', 'REF_RICHER': 'FFEB9C',
        'AI_WRONG': 'FFC7CE', 'REF_WRONG': 'D9D2E9', 'AI_GAVEUP': 'F4CCCC', 'BOTH_UNCLEAR': 'D9D9D9',
    }
    heads = ['Patient', 'Complaint Verdict', 'Diagnosis Verdict', 'REF Diagnosis', 'AI Diagnosis', 'Note']
    for c, h in enumerate(heads, 1):
        cell = ws.cell(1, c, h); cell.fill = HDR; cell.font = Font(color='FFFFFF', bold=True)
    for i, r in enumerate(results, 2):
        ws.cell(i, 1, r['name'])
        c2 = ws.cell(i, 2, r['complaint_verdict']); c2.fill = PatternFill(start_color=colors.get(r['complaint_verdict'],'FFFFFF'), fill_type='solid')
        c3 = ws.cell(i, 3, r['diagnosis_verdict']); c3.fill = PatternFill(start_color=colors.get(r['diagnosis_verdict'],'FFFFFF'), fill_type='solid')
        ws.cell(i, 4, r.get('ref_d',''))
        ws.cell(i, 5, r.get('ai_d',''))
        ws.cell(i, 6, r.get('note',''))
        for col in range(1, 7):
            ws.cell(i, col).alignment = Alignment(wrap_text=True, vertical='top')
    for col, w in zip('ABCDEF', [22, 18, 18, 40, 40, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    wb.save(OUT)

    print('\n' + '='*55)
    print('COMPLAINT verdicts:')
    for k, n in cv.most_common(): print(f'  {k:<14} {n}')
    print('DIAGNOSIS verdicts:')
    for k, n in dv.most_common(): print(f'  {k:<14} {n}')
    print(f'\nReport -> {OUT}')

if __name__ == '__main__':
    asyncio.run(main())
